# -*- coding: utf-8 -*-
"""知识点智能导入服务：Excel 模板 / 文本规则 / AI 抽取 三种模式。

统一流程：上传/粘贴 -> 文本抽取 -> 层级识别 -> 标准化 JSON -> 树形预览确认 -> 事务入库
"""
import hashlib
import io
import json
import re
import uuid
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.exam_detail import KnowledgeSource, SubjectKnowledgePoint
from app.models.subject import Subject


class KnowledgeOcrAdapter:
    """扫描版 OCR 适配器接口（本期占位实现，返回 None 表示暂不支持）。"""

    def extract_text(self, pdf_bytes: bytes) -> Optional[str]:
        return None


def _extract_text_from_bytes(filename: str, content: bytes) -> str:
    """按扩展名抽取文本层；PDF 无文本层时走 OCR 适配器。"""
    name = (filename or "").lower()
    if name.endswith((".txt", ".md", ".csv")):
        return content.decode("utf-8", errors="replace")
    if name.endswith(".docx"):
        from docx import Document
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if name.endswith(".pdf"):
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            if text.strip():
                return text
        except Exception:
            pass
        return KnowledgeOcrAdapter().extract_text(content) or ""
    if name.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if vals:
                    lines.append("\t".join(vals))
        return "\n".join(lines)
    return content.decode("utf-8", errors="replace")


# ---------- 层级识别规则引擎（模式 B） ----------

# 编号体系优先级（类索引）：章节词 < 中文序号 < 括号中文 < 数字层级 < 括号数字
# group(1) 捕获完整编号前缀（含标点），group(2) 为标题
_PATTERNS = [
    # 章节词：第一章 / 第1课 / Unit 1 / Module 1 （体系 0）
    (0, re.compile(r"^(第[0-9一二三四五六七八九十百]+[章节课篇])\s*[、.:：]?\s*(.*)$")),
    (0, re.compile(r"^((?:[Uu]nit|[Mm]odule)\s*\d+)\s*[:：]?\s*(.*)$")),
    # 中文序号：一、二、… （体系 1）
    (1, re.compile(r"^([一二三四五六七八九十百]+、)\s*(.*)$")),
    # 括号中文序号：（一）（二） （体系 2）
    (2, re.compile(r"^((?:（[一二三四五六七八九十百]+）))\s*(.*)$")),
    # 数字层级：1. 1.1 1.1.1 （体系 3/4/5）
    (None, re.compile(r"^((?:\d+(?:\.\d+){0,3})[.、])\s*(.*)$")),
    # 括号数字 / 圈号序号 （体系 6）
    (6, re.compile(r"^([（(]\d+[）)]|[①②③④⑤⑥⑦⑧⑨⑩])\s*(.*)$")),
]


def _join_title(g1: str, g2: str) -> str:
    if not g2:
        return g1
    if g1[-1] in "、）)":
        return g1 + g2
    return f"{g1} {g2}"


def _match_class(line: str):
    """返回 (体系类索引, 标题)。类索引越大层级越深。"""
    for klass, pattern in _PATTERNS:
        m = pattern.match(line)
        if not m:
            continue
        g1 = m.group(1).strip()
        g2 = m.group(2).strip() if len(m.groups()) > 1 else ""
        if klass is None:
            dots = g1.count(".")
            klass = min(dots + 3, 5)  # 1. -> 3 ; 1.1 -> 4 ; 1.1.1 -> 5
        return klass, _join_title(g1, g2)
    return None


def parse_text(text: str) -> List[Dict[str, Any]]:
    """目录/课标文本 -> 层级树（标准化 JSON）。

    编号体系转换：第一章(L1) -> 一、(L2) -> 1.(L3) -> 1.1(L4) -> （1）(L6)。
    """
    roots: List[Dict[str, Any]] = []
    stack: List[Dict[str, Any]] = []  # (klass, node)
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        matched = _match_class(line)
        if not matched:
            # 无法判定的行并入上一层作为描述
            if stack:
                desc = stack[-1][1].get("description") or ""
                stack[-1][1]["description"] = (desc + "\n" + line).strip()
            continue
        klass, title = matched
        node: Dict[str, Any] = {"name": title, "level": klass + 1, "children": [], "description": ""}
        while stack and stack[-1][0] >= klass:
            stack.pop()
        if stack:
            stack[-1][1]["children"].append(node)
        else:
            roots.append(node)
        stack.append((klass, node))
    return roots


def _collect_flat(items: List[Dict[str, Any]], path: str = "") -> List[Dict[str, Any]]:
    """树 -> 扁平列表（含路径，用于预览/去重）。"""
    out = []
    for item in items:
        name = item.get("name", "").strip()
        current_path = f"{path}/{name}" if path else name
        out.append({
            "name": name,
            "level": item.get("level", 1),
            "path": current_path,
            "description": item.get("description", ""),
            "children": item.get("children", []),
        })
        out.extend(_collect_flat(item.get("children", []), current_path))
    return out


def _mark_preview_nodes(items, existing, seen, parent_path: str = ""):
    """在嵌套树上标记 exists / duplicate_in_batch / merge。"""
    out = []
    for item in items:
        name = str(item.get("name", "")).strip()
        path = f"{parent_path}/{name}" if parent_path else name
        dup = seen.get(path, 0)
        seen[path] = dup + 1
        node = {
            "name": name,
            "level": item.get("level", 1),
            "path": path,
            "description": item.get("description", ""),
            "children": _mark_preview_nodes(item.get("children") or [], existing, seen, path),
            "exists": path in existing,
            "duplicate_in_batch": dup > 0,
            "merge": path in existing or dup > 0,
        }
        out.append(node)
    return out


# ---------- Excel 模板（模式 A） ----------

def parse_excel(content: bytes) -> List[Dict[str, Any]]:
    """Excel 模板：列即层级（学段/册/单元/课/节/知识点）。"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if any(vals):
            rows.append(vals)
    if not rows:
        return []
    body = rows[1:]
    roots: List[Dict[str, Any]] = []
    level_node: Dict[int, Dict[str, Any]] = {}
    for vals in body:
        filled = [(i, str(v).strip()) for i, v in enumerate(vals) if v is not None and str(v).strip()]
        if not filled:
            continue
        for i, v in filled:
            node: Dict[str, Any] = {"name": v, "level": i + 1, "children": [], "description": ""}
            # 父级 = 最近一个有内容的低层级列（空列继承上层）
            parent = None
            for j in range(i - 1, -1, -1):
                if j in level_node:
                    parent = level_node[j]
                    break
            if parent:
                parent["children"].append(node)
            else:
                roots.append(node)
            # 该列出现新值，作废更深层级缓存
            for k in list(level_node.keys()):
                if k > i:
                    del level_node[k]
            level_node[i] = node
    return roots


# ---------- AI 抽取（模式 C） ----------

_AI_PROMPT = """你是一个教材目录解析专家。请把下面的教材/课程标准文本解析为知识点层级树，输出严格的 JSON：
{"items":[{"name":"...","level":1,"children":[{"name":"...","level":2}]}]}
要求：
1. 只输出 JSON，不要多余文字；
2. 保持层级准确，level 从 1 开始逐级递增；
3. 标题去重，去掉页码等杂质；
4. 若文本不是目录类内容，items 返回空数组。

文本如下：
{text}"""


def _call_llm_json(text: str) -> List[Dict[str, Any]]:
    """调用 LLM 抽取层级 JSON；失败返回空列表。"""
    if not settings.LLM_ENABLED or not settings.LLM_API_KEY:
        return []
    import httpx
    prompt = _AI_PROMPT.format(text=text[:6000])
    for attempt in range(2):
        try:
            with httpx.Client(timeout=90.0) as client:
                resp = client.post(
                    f"{settings.LLM_API_BASE}/chat/completions",
                    headers={"Authorization": f"Bearer {settings.LLM_API_KEY}"},
                    json={
                        "model": settings.LLM_MODEL,
                        "messages": [
                            {"role": "system", "content": "你是严格输出 JSON 的解析器。"},
                            {"role": "user", "content": prompt},
                        ],
                        "temperature": 0.1,
                    },
                )
                resp.raise_for_status()
                content = resp.json()["choices"][0]["message"]["content"]
                content = content.strip().strip("```json").strip("```").strip()
                data = json.loads(content)
                items = data.get("items", []) if isinstance(data, dict) else data
                if isinstance(items, list) and items:
                    return _normalize_ai_items(items)
        except Exception:
            continue
    return []


def _normalize_ai_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """校验/规整 AI 输出。"""

    def walk(item: Dict[str, Any], level: int) -> Dict[str, Any]:
        name = str(item.get("name", "")).strip()
        if not name:
            return None
        node: Dict[str, Any] = {
            "name": name,
            "level": int(item.get("level", level) or level),
            "children": [],
            "description": "",
        }
        children = item.get("children") or []
        for c in children:
            child = walk(c, node["level"] + 1)
            if child:
                node["children"].append(child)
        return node

    result = []
    for item in items:
        node = walk(item, 1)
        if node:
            result.append(node)
    return result


def parse_ai(text: str) -> List[Dict[str, Any]]:
    """AI 抽取，失败回退规则引擎。"""
    items = _call_llm_json(text)
    if items:
        return items
    return parse_text(text)


# ---------- 预览与入库 ----------

def _existing_paths(db: Session, subject_id: uuid.UUID) -> Dict[str, uuid.UUID]:
    """现有知识点路径 -> id 映射（路径以 / 分隔）。"""
    kps = db.query(SubjectKnowledgePoint).filter(SubjectKnowledgePoint.subject_id == subject_id).all()
    by_id = {kp.id: kp for kp in kps}
    path_map: Dict[str, uuid.UUID] = {}
    for kp in kps:
        parts = []
        cur = kp
        while cur:
            parts.append(cur.name)
            cur = by_id.get(cur.parent_id) if cur.parent_id else None
        path = "/".join(reversed(parts))
        path_map[path] = kp.id
    return path_map


def build_preview(db: Session, subject_id: uuid.UUID, items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """构建预览：嵌套树节点标记 exists / duplicate，确认后原样提交入库。"""
    existing = _existing_paths(db, subject_id)
    seen: Dict[str, int] = {}
    preview_items = _mark_preview_nodes(items, existing, seen)

    def _count(pred, nodes):
        total = 0
        for n in nodes:
            if pred(n):
                total += 1
            total += _count(pred, n.get("children") or [])
        return total

    return {
        "subject_id": str(subject_id),
        "item_count": _count_items(preview_items),
        "exists_count": _count(lambda n: n["exists"], preview_items),
        "duplicate_count": _count(lambda n: n["duplicate_in_batch"], preview_items),
        "items": preview_items,
    }


def commit_import(
    db: Session,
    subject_id: uuid.UUID,
    source_name: str,
    source_type: str,
    import_mode: str,
    items: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """事务入库：同名合并（父同名->合并追加子节点；叶子同名->合并标记）。"""
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        raise ValueError("科目不存在")
    source = KnowledgeSource(
        subject_id=subject_id,
        source_name=source_name or f"{subject.name} 知识点导入",
        source_type=source_type,
        import_mode=import_mode,
        status="imported",
        meta={"item_count": _count_items(items)},
    )
    db.add(source)
    db.flush()

    created = 0
    merged = 0

    def insert_tree(node: Dict[str, Any], parent_id: Optional[uuid.UUID], sort_index: int) -> None:
        nonlocal created, merged
        name = str(node.get("name", "")).strip()
        if not name:
            return
        existing = (
            db.query(SubjectKnowledgePoint)
            .filter(
                SubjectKnowledgePoint.subject_id == subject_id,
                SubjectKnowledgePoint.name == name,
                SubjectKnowledgePoint.parent_id == parent_id,
            )
            .first()
        )
        if existing:
            merged += 1
            kp = existing
        else:
            kp = SubjectKnowledgePoint(
                subject_id=subject_id,
                name=name,
                parent_id=parent_id,
                sort_order=sort_index,
                description=node.get("description", ""),
                source_id=source.id,
                origin="imported",
            )
            db.add(kp)
            db.flush()
            created += 1
        for i, child in enumerate(node.get("children") or []):
            insert_tree(child, kp.id, i + 1)

    for i, root in enumerate(items):
        insert_tree(root, None, i + 1)
    db.commit()
    return {"source_id": str(source.id), "created": created, "merged": merged, "total": created + merged}


def _count_items(items: List[Dict[str, Any]]) -> int:
    total = 0
    for item in items:
        total += 1 + _count_items(item.get("children") or [])
    return total
