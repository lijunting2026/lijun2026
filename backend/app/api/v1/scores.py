from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import io, uuid, urllib.parse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.database import get_db
from app.models.score import Score
from app.models.student import Student
from app.models.exam import Exam, ExamSubject
from app.models.subject import Subject
from app.models.school import ClassInfo
from sqlalchemy.orm import aliased
from app.schemas.score import ScoreCreate, ScoreBatchCreate, ScoreUpdate, ScoreResponse


router = APIRouter(prefix="/scores", tags=["成绩管理"])


def _try_uuid(value: str, name: str = "ID") -> uuid.UUID:
    try:
        return uuid.UUID(value)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail=f"无效的{name}")


def _get_student_current_class_id(db: Session, student_id: uuid.UUID) -> uuid.UUID:
    student = db.query(Student.class_id).filter(Student.id == student_id).first()
    return student[0] if student else None


@router.get("/")
def list_scores(
    exam_id: str = None, class_id: str = None, grade_id: str = None,
    date_from: str = None, date_to: str = None,
    skip: int = 0, limit: int = 1000,
    db: Session = Depends(get_db),
):
    q = db.query(Score)
    needs_exam_subject = exam_id is not None or grade_id is not None or date_from or date_to
    needs_exam = grade_id is not None or date_from or date_to
    needs_class = class_id is not None

    if needs_exam_subject:
        q = q.join(Score.exam_subject)
    if needs_exam:
        q = q.join(ExamSubject.exam)
    if needs_class:
        q = q.join(Score.class_info)

    if exam_id:
        q = q.filter(ExamSubject.exam_id == _try_uuid(exam_id, "考试ID"))
    if grade_id:
        q = q.filter(Exam.grade_id == _try_uuid(grade_id, "年级ID"))
    if date_from:
        q = q.filter(Exam.exam_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        q = q.filter(Exam.exam_date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    if class_id:
        q = q.filter(ClassInfo.id == _try_uuid(class_id, "班级ID"))

    total = q.count()
    scores = q.offset(skip).limit(limit).all()
    result = []
    for s in scores:
        student = db.query(Student).filter(Student.id == s.student_id).first()
        subj = s.exam_subject.subject if s.exam_subject else None
        cls = s.class_info if s.class_id else None
        result.append({
            "id": str(s.id),
            "student_id": str(s.student_id),
            "student_no": student.student_no if student else None,
            "student_name": student.name if student else None,
            "exam_subject_id": str(s.exam_subject_id),
            "subject_name": subj.name if subj else None,
            "score_value": s.score_value,
            "status": s.status,
            "class_id": str(s.class_id) if s.class_id else None,
            "class_name": cls.name if cls else None,
            "exam_id": str(s.exam_subject.exam.id) if s.exam_subject and s.exam_subject.exam else None,
            "exam_name": s.exam_subject.exam.name if s.exam_subject and s.exam_subject.exam else None,
        })
    return {"total": total, "items": result}


@router.post("/batch")
def batch_create_scores(data: ScoreBatchCreate, db: Session = Depends(get_db)):
    """批量录入/更新成绩。
    新增记录：class_id 记录学生当前班级（即考试时的班级归属）。
    更新已有记录：保留原有的 class_id（成绩归属不随转班改变）。
    """
    count = 0
    for s in data.scores:
        sid = uuid.UUID(s.student_id)
        esid = uuid.UUID(s.exam_subject_id)
        student_class_id = _get_student_current_class_id(db, sid)

        existing = (
            db.query(Score)
            .filter(Score.student_id == sid, Score.exam_subject_id == esid)
            .first()
        )
        if existing:
            existing.score_value = s.score_value
            existing.status = s.status
            # 注意：不更新 class_id —— 成绩归属考试时的班级，不随转班改变
        else:
            db.add(
                Score(
                    student_id=sid,
                    exam_subject_id=esid,
                    score_value=s.score_value,
                    status=s.status,
                    class_id=student_class_id,
                )
            )
        count += 1
    db.commit()
    return {"message": "成功保存 %d 条成绩" % count}



from pydantic import BaseModel

class BatchDeleteRequest(BaseModel):
    ids: List[str]


@router.post("/batch-delete")
def batch_delete_scores(data: BatchDeleteRequest, db: Session = Depends(get_db)):
    deleted = 0
    for sid in data.ids:
        score = db.query(Score).filter(Score.id == _try_uuid(sid, "成绩ID")).first()
        if score:
            db.delete(score)
            deleted += 1
    db.commit()
    return {"message": "成功删除 %d 条成绩" % deleted, "deleted": deleted}


@router.get("/export-template")
def export_template(
    exam_id: str,
    subject_id: Optional[str] = None,
    class_id: Optional[str] = None,
    db: Session = Depends(get_db),
):
    exam_uuid = _try_uuid(exam_id, "考试ID")
    exam = db.query(Exam).filter(Exam.id == exam_uuid).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    eq = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_uuid)
    if subject_id:
        eq = eq.filter(ExamSubject.subject_id == _try_uuid(subject_id, "科目ID"))
    exam_subjects = eq.all()
    if not exam_subjects:
        raise HTTPException(status_code=400, detail="该考试没有科目数据")

    sq = db.query(Student)
    if class_id:
        sq = sq.filter(Student.class_id == _try_uuid(class_id, "班级ID"))
    else:
        cids = [c[0] for c in db.query(ClassInfo.id).filter(ClassInfo.grade_id == exam.grade_id).all()]
        if cids:
            sq = sq.filter(Student.class_id.in_(cids))
    students = sq.order_by(Student.student_no).all()

    subjects = db.query(Subject).all()
    subj_map = {s.id: s for s in subjects}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩模板"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["学号", "姓名"]
    for es in exam_subjects:
        subj = subj_map.get(es.subject_id)
        headers.append("%s\n(满分%d)" % (subj.name if subj else "未知", es.full_score))

    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for s in students:
        row = [s.student_no, s.name]
        for es in exam_subjects:
            row.append("")
        ws.append(row)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    for ci in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + urllib.parse.quote("%s_成绩模板.xlsx" % exam.name)},
    )


@router.post("/import")
async def import_scores(
    exam_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """从 Excel 导入成绩。
    新增记录：class_id 记录学生当前班级。
    更新已有记录：保留原有的 class_id，不随转班改变。
    """
    exam_uuid = _try_uuid(exam_id, "考试ID")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="文件没有数据行")

    cols = [str(h).strip() if h else "" for h in rows[0]]
    sno_col = None
    subj_cols = {}
    for i, h in enumerate(cols):
        hc = h.split("\n")[0].strip()
        if hc == "学号":
            sno_col = i
        elif hc != "姓名":
            subj_cols[i] = hc

    if sno_col is None:
        raise HTTPException(status_code=400, detail="缺少「学号」列")

    exam_subjects = db.query(ExamSubject).filter(ExamSubject.exam_id == exam_uuid).all()
    all_subs = db.query(Subject).all()
    name_to_id = {s.name: s.id for s in all_subs}
    es_map = {es.subject_id: es for es in exam_subjects}

    imported = 0
    errors = []
    for ri, row in enumerate(rows[1:], start=2):
        sno = str(row[sno_col]).strip() if row[sno_col] is not None else ""
        if not sno:
            continue
        student = db.query(Student).filter(Student.student_no == sno).first()
        if not student:
            errors.append("行%d: 学号%s不存在" % (ri, sno))
            continue
        student_class_id = student.class_id
        for ci, sn in subj_cols.items():
            sv = row[ci]
            if sv is None or str(sv).strip() == "":
                continue
            try:
                sv = float(sv)
            except (ValueError, TypeError):
                errors.append("行%d: %s分数无效" % (ri, sn))
                continue
            sid = name_to_id.get(sn)
            if not sid:
                errors.append("行%d: 科目%s不存在" % (ri, sn))
                continue
            es = es_map.get(sid)
            if not es:
                errors.append("行%d: 科目%s不属于该考试" % (ri, sn))
                continue
            existing = (
                db.query(Score)
                .filter(Score.student_id == student.id, Score.exam_subject_id == es.id)
                .first()
            )
            if existing:
                existing.score_value = sv
                # 注意：不更新 class_id —— 保留原有班级归属
            else:
                db.add(
                    Score(
                        student_id=student.id,
                        exam_subject_id=es.id,
                        score_value=sv,
                        class_id=student_class_id,
                    )
                )
            imported += 1

    db.commit()
    msg = "成功导入%d条成绩" % imported
    if errors:
        msg += "，%d条错误" % len(errors)
    return {"message": msg, "imported": imported, "errors": errors[:20]}


@router.get("/export")
def export_scores(exam_id: str, db: Session = Depends(get_db)):
    return {"data": list_scores(exam_id=exam_id, db=db), "filename": "scores_%s.xlsx" % exam_id}


@router.put("/{score_id}")
def update_score(score_id: str, data: ScoreUpdate, db: Session = Depends(get_db)):
    score = db.query(Score).filter(Score.id == _try_uuid(score_id, "成绩ID")).first()
    if not score:
        raise HTTPException(status_code=404, detail="成绩记录不存在")
    if data.score_value is not None:
        score.score_value = data.score_value
    if data.status is not None:
        score.status = data.status
    db.commit()
    db.refresh(score)
    student = db.query(Student).filter(Student.id == score.student_id).first()
    subj = score.exam_subject.subject if score.exam_subject else None
    return {
        "id": str(score.id),
        "student_id": str(score.student_id),
        "student_no": student.student_no if student else None,
        "student_name": student.name if student else None,
        "exam_subject_id": str(score.exam_subject_id),
        "subject_name": subj.name if subj else None,
        "score_value": score.score_value,
        "status": score.status,
    }


@router.delete("/{score_id}")
def delete_score(score_id: str, db: Session = Depends(get_db)):
    score = db.query(Score).filter(Score.id == _try_uuid(score_id, "成绩ID")).first()
    if not score:
        raise HTTPException(status_code=404, detail="成绩记录不存在")
    db.delete(score)
    db.commit()
    return {"message": "已删除"}


@router.get("/summary")
def score_summary(
    exam_id: str = None,
    grade_id: str = None,
    class_id: str = None,
    subject_id: str = None,
    db: Session = Depends(get_db),
):
    """成绩聚合统计：按考试/年级/班级汇总平均分、及格率、优秀率等"""
    from sqlalchemy import func
    import statistics

    q = db.query(Score)
    q = q.join(Score.exam_subject).join(ExamSubject.exam).join(ExamSubject.subject)
    q = q.join(Exam.grade)

    if exam_id:
        q = q.filter(Exam.id == _try_uuid(exam_id, "考试ID"))
    if grade_id:
        q = q.filter(Exam.grade_id == _try_uuid(grade_id, "年级ID"))
    if class_id:
        q = q.filter(Score.class_id == _try_uuid(class_id, "班级ID"))
    if subject_id:
        q = q.filter(ExamSubject.subject_id == _try_uuid(subject_id, "科目ID"))

    scores = q.all()
    if not scores:
        return {
            "exam_name": "", "grade_name": "", "total_students": 0,
            "total_subjects": 0, "subject_summaries": [], "class_summaries": [],
            "overall_avg": 0, "overall_pass_rate": 0
        }

    exam = scores[0].exam_subject.exam
    exam_name = exam.name
    grade_name = exam.grade.name if exam.grade else ""

    # Group by subject
    from collections import defaultdict
    subj_data = defaultdict(list)
    class_data = defaultdict(list)
    all_students = set()

    for s in scores:
        subj_name = s.exam_subject.subject.name
        subj_id = str(s.exam_subject.subject_id)
        full_score = s.exam_subject.full_score
        subj_data[(subj_id, subj_name, full_score)].append(s.score_value)
        if s.class_id:
            cls = s.class_info
            cls_name = cls.name if cls else "未知"
            class_data[(str(s.class_id), cls_name)].append(s.score_value)
        all_students.add(str(s.student_id))

    # Subject summaries
    subject_summaries = []
    for (sid, sname, full), vals in subj_data.items():
        if not vals:
            continue
        sorted_vals = sorted(vals)
        n = len(sorted_vals)
        pass_count = sum(1 for v in vals if v >= full * 0.6)
        excellence_count = sum(1 for v in vals if v >= full * 0.85)
        fail_count = sum(1 for v in vals if v < full * 0.6)
        median_val = sorted_vals[n // 2] if n % 2 else (sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2
        mean_val = sum(vals) / n
        variance = sum((v - mean_val) ** 2 for v in vals) / n if n > 1 else 0

        subject_summaries.append({
            "subject_id": sid,
            "subject_name": sname,
            "full_score": full,
            "avg_score": round(mean_val, 1),
            "max_score": max(vals),
            "min_score": min(vals),
            "median_score": round(median_val, 1),
            "pass_count": pass_count,
            "total_count": n,
            "pass_rate": round(pass_count / n * 100, 1),
            "excellence_count": excellence_count,
            "excellence_rate": round(excellence_count / n * 100, 1),
            "fail_count": fail_count,
            "fail_rate": round(fail_count / n * 100, 1),
            "std_dev": round(variance ** 0.5, 1),
        })

    # Class summaries
    class_summaries = []
    for (cid, cname), vals in class_data.items():
        n = len(vals)
        if n == 0:
            continue
        class_summaries.append({
            "class_id": cid,
            "class_name": cname,
            "avg_score": round(sum(vals) / n, 1),
            "max_score": max(vals),
            "min_score": min(vals),
            "total_count": n,
            "rank": 0,
        })
    # Rank classes by avg_score
    class_summaries.sort(key=lambda x: x["avg_score"], reverse=True)
    for i, c in enumerate(class_summaries):
        c["rank"] = i + 1

    # Overall stats
    all_vals = [s.score_value for s in scores]
    overall_avg = round(sum(all_vals) / len(all_vals), 1) if all_vals else 0
    full_scores_map = {str(s.exam_subject.subject_id): s.exam_subject.full_score for s in scores}
    overall_pass = 0
    student_subjects = defaultdict(list)
    for s in scores:
        student_subjects[str(s.student_id)].append((s.score_value, full_scores_map.get(str(s.exam_subject.subject_id), 100)))
    for sid, subj_scores in student_subjects.items():
        avg_pct = sum(v / fs for v, fs in subj_scores) / len(subj_scores) if subj_scores else 0
        if avg_pct >= 0.6:
            overall_pass += 1
    total_students = len(all_students)
    overall_pass_rate = round(overall_pass / total_students * 100, 1) if total_students else 0

    return {
        "exam_name": exam_name,
        "grade_name": grade_name,
        "total_students": total_students,
        "total_subjects": len(subject_summaries),
        "subject_summaries": subject_summaries,
        "class_summaries": class_summaries,
        "overall_avg": overall_avg,
        "overall_pass_rate": overall_pass_rate,
    }
