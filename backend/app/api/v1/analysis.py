from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.schemas.analysis import ExamAnalysisResponse, DistributionResponse, DashboardResponse
from app.services.analytics.analysis_service import AnalysisService
from app.services.analytics.dashboard_service import DashboardService
from app.services.analytics.student_tracking import StudentTrackingService, ClassAnalysisService


router = APIRouter(prefix="/analysis", tags=["数据分析"])


@router.get("/exam/{exam_id}", response_model=ExamAnalysisResponse)
def analyze_exam(exam_id: str, db: Session = Depends(get_db)):
    service = AnalysisService(db)
    return service.get_exam_analysis(exam_id)


@router.get("/distribution/{exam_subject_id}", response_model=DistributionResponse)
def get_distribution(exam_subject_id: str, bins: int = 10, db: Session = Depends(get_db)):
    service = AnalysisService(db)
    return service.get_score_distribution(exam_subject_id, bins)


@router.get("/student/{student_id}")
def get_student_analysis(student_id: str, db: Session = Depends(get_db)):
    try:
        service = StudentTrackingService(db)
        result = service.get_student_scores(student_id)
        if not result:
            raise HTTPException(status_code=404, detail="学生不存在")
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/student/{student_id}/advice")
def get_student_advice(student_id: str, db: Session = Depends(get_db)):
    try:
        service = StudentTrackingService(db)
        result = service.generate_advice(student_id)
        if not result:
            raise HTTPException(status_code=404, detail="学生不存在")
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/class/{class_id}")
def get_class_analysis(class_id: str, db: Session = Depends(get_db)):
    try:
        service = ClassAnalysisService(db)
        result = service.get_class_overview(class_id)
        if not result:
            raise HTTPException(status_code=404, detail="班级不存在")
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/class/{class_id}/export")
def export_class_analysis(class_id: str, db: Session = Depends(get_db)):
    import io, urllib.parse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    service = ClassAnalysisService(db)
    data = service.get_class_overview(class_id)
    if not data:
        raise HTTPException(status_code=404, detail="班级不存在")

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Sheet 1: 班级概况
    ws = wb.active
    ws.title = "班级概况"
    ws.append(["班级", data.get("class_name", "")])
    ws.append(["年级", data.get("grade_name", "")])
    ws.append(["学生人数", data.get("student_count", 0)])
    ws.append(["考试次数", data.get("exam_count", 0)])
    ws.append([])

    headers = ["科目", "平均分", "最高分", "最低分", "样本数"]
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(row=6, column=ci)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for ss in (data.get("subject_stats") or []):
        ws.append([ss["subject_name"], ss["avg_score"], ss["max_score"], ss["min_score"], ss["count"]])

    ws.column_dimensions["A"].width = 14
    for col in "BCDE":
        ws.column_dimensions[col].width = 12

    # Sheet 2: 考试成绩趋势
    ws2 = wb.create_sheet("考试趋势")
    ws2.append(["考试名称", "考试日期", "平均得分率(%)"])
    for ci in range(1, 4):
        c2 = ws2.cell(row=1, column=ci)
        c2.font, c2.fill, c2.alignment, c2.border = hf, hfill, ha, tb
    for es in (data.get("exam_summary") or []):
        ws2.append([es["exam_name"], es.get("exam_date", ""), es.get("avg_rate", "")])

    # Sheet 3: 学生列表
    ws3 = wb.create_sheet("学生列表")
    ws3.append(["学号", "姓名", "综合得分率(%)"])
    for ci in range(1, 4):
        c3 = ws3.cell(row=1, column=ci)
        c3.font, c3.fill, c3.alignment, c3.border = hf, hfill, ha, tb
    for sl in (data.get("student_list") or []):
        ws3.append([sl["student_no"], sl["student_name"], sl.get("avg_rate", "")])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    name = "%s_班级分析.xlsx" % data.get("class_name", "班级")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + urllib.parse.quote(name)},
    )


@router.get("/student/{student_id}/export")
def export_student_analysis(student_id: str, db: Session = Depends(get_db)):
    import io, urllib.parse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    service = StudentTrackingService(db)
    data = service.get_student_scores(student_id)
    if not data:
        raise HTTPException(status_code=404, detail="学生不存在")

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Sheet 1: 学生概况
    ws = wb.active
    ws.title = "学情概况"
    ws.append(["学号", data.get("student_no", "")])
    ws.append(["姓名", data.get("student_name", "")])
    ws.append(["班级", data.get("class_name", "")])
    ws.append(["年级", data.get("grade_name", "")])
    ws.append(["考试次数", data.get("exam_count", 0)])
    ws.append(["总体趋势", data.get("overall_trend", "")])
    ws.append([])

    # Sheet 2: 历次成绩
    ws2 = wb.create_sheet("历次成绩")
    ws2.append(["考试名称", "考试日期", "综合得分率(%)", "科目", "得分", "得分率(%)"])
    for ci in range(1, 7):
        c2 = ws2.cell(row=1, column=ci)
        c2.font, c2.fill, c2.alignment, c2.border = hf, hfill, ha, tb
    for exam in (data.get("exams") or []):
        ws2.append([exam["exam_name"], exam.get("exam_date", ""), exam.get("avg_rate", ""), "", "", ""])
        for subj in (exam.get("subjects") or []):
            ws2.append(["", "", "", subj.get("subject_name", ""), subj.get("score", ""), subj.get("rate", "")])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 12

    # Sheet 3: 各科趋势
    ws3 = wb.create_sheet("各科趋势")
    headers3 = ["科目", "趋势描述", "历次得分率"]
    ws3.append(headers3)
    for ci in range(1, 4):
        c3 = ws3.cell(row=1, column=ci)
        c3.font, c3.fill, c3.alignment, c3.border = hf, hfill, ha, tb
    for t in (data.get("trends") or []):
        scores_str = "; ".join(["%s: %s%%" % (s.get("exam_name", ""), s.get("rate", "")) for s in (t.get("scores") or [])])
        ws3.append([t.get("subject_name", ""), t.get("description", ""), scores_str])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    name = "%s_学情分析.xlsx" % data.get("student_name", "学生")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + urllib.parse.quote(name)},
    )


@router.get("/dashboard", response_model=DashboardResponse)
def get_dashboard(db: Session = Depends(get_db)):
    try:
        service = DashboardService(db)
        return service.get_dashboard()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/{exam_id}")
def export_analysis(exam_id: str, db: Session = Depends(get_db)):
    import io, urllib.parse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    service = AnalysisService(db)
    data = service.get_exam_analysis(exam_id)
    if not data:
        raise HTTPException(status_code=404, detail="考试不存在")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分析结果"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # 表头
    headers = ["科目", "平均分", "得分率(%)", "最高分", "最低分", "及格率(%)", "优秀率(%)", "标准差"]
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    # 数据行
    for gs in (data.get("grade_stats") or []):
        ws.append([gs["subject_name"], gs["avg_score"], gs["avg_score_rate"], gs["max_score"],
                    gs["min_score"], gs["pass_rate"], gs["excellent_rate"], gs["std_dev"]])

    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 12

    # 班级统计表 - 第二个sheet
    ws2 = wb.create_sheet("各班统计")
    ws2.append(["班级", "人数", "科目", "平均分", "得分率(%)", "及格率(%)", "优秀率(%)"])
    for ci in range(1, 8):
        c2 = ws2.cell(row=1, column=ci)
        c2.font, c2.fill, c2.alignment, c2.border = hf, hfill, ha, tb
    for cs in (data.get("class_stats") or []):
        for s in (cs.get("stats") or []):
            ws2.append([cs["class_name"], cs["student_count"], s["subject_name"],
                        s["avg_score"], s["avg_score_rate"], s["pass_rate"], s["excellent_rate"]])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    name = "%s_分析结果.xlsx" % data.get("exam_name", "考试")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + urllib.parse.quote(name)},
    )


@router.post("/chat")
def ai_chat(data: dict, db: Session = Depends(get_db)):
    from app.services.ai.chat_service import AIChatService
    try:
        message = data.get("message", "")
        context_type = data.get("context_type", "general")
        context_id = data.get("context_id")
        service = AIChatService(db)
        result = service.chat(message, context_type, context_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
