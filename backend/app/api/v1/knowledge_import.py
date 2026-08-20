# -*- coding: utf-8 -*-
"""知识点智能导入 API：模板下载 / Excel / 文本 / AI / 预览确认 / 来源管理。"""
import io
import uuid
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.exam_detail import KnowledgeSource, SubjectKnowledgePoint
from app.models.subject import Subject
from app.services.analytics.knowledge_import_service import (
    build_preview,
    commit_import,
    parse_ai,
    parse_excel,
    parse_text,
    _extract_text_from_bytes,
)

router = APIRouter(prefix="/knowledge-points", tags=["知识点导入"])


class TextImportRequest(BaseModel):
    subject_id: str
    source_name: str = ""
    text: str
    source_type: str = "textbook"


class PreviewCommitRequest(BaseModel):
    subject_id: str
    source_name: str = ""
    source_type: str = "textbook"
    import_mode: str = "rules"
    items: list


def _get_subject(db: Session, subject_id: str) -> Subject:
    try:
        subj = db.query(Subject).filter(Subject.id == uuid.UUID(subject_id)).first()
    except ValueError:
        raise HTTPException(status_code=400, detail="无效的科目ID")
    if not subj:
        raise HTTPException(status_code=404, detail="科目不存在")
    return subj


def _preview_response(db: Session, subject_id: str, items: list):
    preview = build_preview(db, uuid.UUID(subject_id), items)
    if not preview["items"]:
        raise HTTPException(status_code=400, detail="未能识别出任何知识点，请检查输入内容")
    return preview


@router.get("/import/template.xlsx")
def download_import_template():
    """下载模式 A 模板（列=层级）。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "知识点导入模板"
    headers = ["学段", "册次", "单元", "课", "节", "知识点"]
    ws.append(headers)
    ws.append(["高中", "必修第一册", "第一章 集合", "1.1 集合的概念", "", "集合中元素的特性"])
    ws.append(["高中", "必修第一册", "第一章 集合", "1.1 集合的概念", "", "元素与集合的关系"])
    ws.append(["高中", "必修第一册", "第一章 集合", "1.2 集合间的基本关系", "", "子集与真子集"])
    ws.append(["高中", "必修第一册", "第一章 集合", "1.3 集合的基本运算", "", "并集与交集"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        width = max(len(str(cell.value or "")) * 2 + 4 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width, 12), 30)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=knowledge_import_template.xlsx"},
    )


@router.post("/import/excel")
def import_excel(
    file: UploadFile = File(...),
    subject_id: str = "",
    source_name: str = "",
    db: Session = Depends(get_db),
):
    """模式 A：Excel 模板导入 -> 预览数据。"""
    subj = _get_subject(db, subject_id)
    content = file.file.read()
    items = parse_excel(content)
    return _preview_response(db, str(subj.id), items)


@router.post("/import/text")
def import_text(data: TextImportRequest, db: Session = Depends(get_db)):
    """模式 B：目录/课标文本 -> 预览数据。"""
    subj = _get_subject(db, data.subject_id)
    items = parse_text(data.text)
    return _preview_response(db, str(subj.id), items)


@router.post("/import/ai")
def import_ai(
    file: UploadFile = File(...),
    subject_id: str = "",
    source_name: str = "",
    db: Session = Depends(get_db),
):
    """模式 C：Word/PDF/文本 AI 抽取 -> 预览数据（未配置 LLM 自动回退规则解析）。"""
    subj = _get_subject(db, subject_id)
    content = file.file.read()
    text = _extract_text_from_bytes(file.filename or "", content)
    if not text.strip():
        raise HTTPException(status_code=400, detail="未能从文件中提取文本，扫描版暂不支持（OCR 接口预留中）")
    items = parse_ai(text)
    return _preview_response(db, str(subj.id), items)


@router.post("/import/preview")
def commit_preview(data: PreviewCommitRequest, db: Session = Depends(get_db)):
    """提交预览确认结果 -> 事务入库（同名合并）。"""
    subj = _get_subject(db, data.subject_id)
    try:
        result = commit_import(
            db,
            subj.id,
            source_name=data.source_name,
            source_type=data.source_type,
            import_mode=data.import_mode,
            items=data.items,
        )
        return {"message": f"导入完成：新增 {result['created']} 个，合并 {result['merged']} 个", **result}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/sources")
def list_sources(subject_id: str = "", db: Session = Depends(get_db)):
    """导入来源/批次列表。"""
    q = db.query(KnowledgeSource)
    if subject_id:
        q = q.filter(KnowledgeSource.subject_id == uuid.UUID(subject_id))
    sources = q.order_by(KnowledgeSource.created_at.desc()).all()
    return [{
        "id": str(s.id),
        "subject_id": str(s.subject_id),
        "source_name": s.source_name,
        "source_type": s.source_type,
        "import_mode": s.import_mode,
        "status": s.status,
        "meta": s.meta or {},
        "created_at": str(s.created_at) if s.created_at else None,
    } for s in sources]


@router.delete("/sources/{source_id}")
def delete_source(source_id: str, db: Session = Depends(get_db)):
    """删除导入来源批次（同时清理 origin=imported 且 source_id 指向它的知识点）。"""
    source = db.query(KnowledgeSource).filter(KnowledgeSource.id == uuid.UUID(source_id)).first()
    if not source:
        raise HTTPException(status_code=404, detail="来源不存在")
    db.query(SubjectKnowledgePoint).filter(
        SubjectKnowledgePoint.source_id == source.id,
        SubjectKnowledgePoint.origin == "imported",
    ).delete(synchronize_session=False)
    db.delete(source)
    db.commit()
    return {"message": "来源及其导入知识点已删除"}
