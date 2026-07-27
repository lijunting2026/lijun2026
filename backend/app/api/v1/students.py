from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import List
from app.core.database import get_db
from app.models.student import Student
from app.models.score import Score
from app.models.school import ClassInfo
from app.schemas.student import StudentCreate, StudentResponse, StudentImport, TransferRequest, TransferResponse
import io, uuid, urllib.parse
from app.utils import parse_uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/students", tags=["学生管理"])


def _build_student_response(student: Student) -> StudentResponse:
    class_name = student.class_info.name if student.class_info else None
    grade_name = student.class_info.grade.name if student.class_info and student.class_info.grade else None
    return StudentResponse(
        id=str(student.id),
        student_no=student.student_no,
        name=student.name,
        gender=student.gender,
        class_id=str(student.class_id),
        class_name=class_name,
        grade_name=grade_name,
        created_at=student.created_at,
    )


@router.get("/")
def list_students(
    class_id: str = None, keyword: str = None,
    skip: int = 0, limit: int = 100,
    db: Session = Depends(get_db),
):
    q = db.query(Student)
    if class_id:
        q = q.filter(Student.class_id == parse_uuid(class_id))
    if keyword:
        q = q.filter(
            Student.name.like(f"%{keyword}%")
            | Student.student_no.like(f"%{keyword}%")
        )
    total = q.count()
    students = q.order_by(Student.student_no).offset(skip).limit(limit).all()
    items = [_build_student_response(s) for s in students]
    return {"total": total, "items": items}


@router.post("/", response_model=StudentResponse)
def create_student(data: StudentCreate, db: Session = Depends(get_db)):
    existing = db.query(Student).filter(Student.student_no == data.student_no).first()
    if existing:
        raise HTTPException(status_code=400, detail="学号已存在")
    cls = db.query(ClassInfo).filter(ClassInfo.id == parse_uuid(data.class_id)).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")
    student = Student(
        student_no=data.student_no, name=data.name,
        gender=data.gender, class_id=parse_uuid(data.class_id),
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return _build_student_response(student)


@router.post("/import")
def import_students(data: StudentImport, db: Session = Depends(get_db)):
    count = 0
    for s in data.students:
        existing = db.query(Student).filter(Student.student_no == s.student_no).first()
        if not existing:
            student = Student(
                student_no=s.student_no, name=s.name,
                gender=s.gender, class_id=s.class_id,
            )
            db.add(student)
            count += 1
    db.commit()
    return {"message": f"成功导入 {count} 名学生"}


@router.get("/export-template")
def export_student_template(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生模板"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["学号", "姓名", "性别"]
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=ci)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    ws.append(["202601001", "张三", "男"])

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename="
            + urllib.parse.quote("学生导入模板.xlsx")
        },
    )


@router.post("/import-excel")
async def import_students_excel(
    class_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="文件没有数据行")

    imported = 0
    errors = []
    for ri, row in enumerate(rows[1:], start=2):
        sno = str(row[0]).strip() if row[0] is not None else ""
        name = str(row[1]).strip() if row[1] is not None else ""
        gender = (
            str(row[2]).strip()
            if row[2] is not None and str(row[2]).strip()
            else "未知"
        )
        if not sno or not name:
            errors.append(f"行{ri}: 学号或姓名为空")
            continue
        existing = db.query(Student).filter(Student.student_no == sno).first()
        if existing:
            errors.append(f"行{ri}: 学号{sno}已存在")
            continue
        gender_val = gender if gender in ("男", "女") else "未知"
        student = Student(
            student_no=sno, name=name,
            gender=gender_val, class_id=parse_uuid(class_id),
        )
        db.add(student)
        imported += 1

    db.commit()
    msg = f"成功导入{imported}名学生"
    if errors:
        msg += f"，{len(errors)}条错误"
    return {"message": msg, "imported": imported, "errors": errors[:20]}


@router.put("/{student_id}", response_model=StudentResponse)
def update_student(student_id: str, data: StudentCreate, db: Session = Depends(get_db)):
    student = db.query(Student).filter(Student.id == parse_uuid(student_id)).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    existing = (
        db.query(Student)
        .filter(Student.student_no == data.student_no, Student.id != parse_uuid(student_id))
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="学号已被其他学生使用")

    student.student_no = data.student_no
    student.name = data.name
    student.gender = data.gender or "未知"
    student.class_id = parse_uuid(data.class_id)
    # 注意：不迁移已有成绩的 class_id
    # Score.class_id 记录的是考试时的班级归属，不随转班改变

    db.commit()
    db.refresh(student)
    return _build_student_response(student)


@router.post("/{student_id}/transfer", response_model=TransferResponse)
def transfer_student(
    student_id: str, data: TransferRequest, db: Session = Depends(get_db),
):
    """学生转班接口。
    - 更新 Student.class_id
    - 默认不迁移已有成绩（成绩归属考试时的班级）
    - 当 migrate_scores=True 时，所有历史成绩的 class_id 改为新班级
    """
    student = db.query(Student).filter(Student.id == parse_uuid(student_id)).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    target_class = db.query(ClassInfo).filter(ClassInfo.id == parse_uuid(data.target_class_id)).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="目标班级不存在")

    original_class_name = student.class_info.name if student.class_info else ""

    # 更新学生所在班级
    student.class_id = target_class.id

    # 按需迁移已有成绩的 class_id
    migrated_count = 0
    if data.migrate_scores:
        migrated_count = (
            db.query(Score)
            .filter(Score.student_id == student.id)
            .update({"class_id": target_class.id})
        )

    db.commit()
    db.refresh(student)

    return TransferResponse(
        id=str(student.id),
        student_no=student.student_no,
        student_name=student.name,
        original_class_name=original_class_name,
        target_class_name=target_class.name,
        migrated_score_count=migrated_count or 0,
        scores_follow_student=data.migrate_scores,
    )


@router.delete("/{student_id}")
def delete_student(student_id: str, db: Session = Depends(get_db)):
    student = db.query(Student).filter(Student.id == parse_uuid(student_id)).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    db.delete(student)
    db.commit()
    return {"message": "已删除"}
