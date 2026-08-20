from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import uuid
from app.utils import parse_uuid
from app.core.database import get_db
from app.models.exam import Exam, ExamSubject
from app.models.subject import Subject
from app.schemas.exam import ExamCreate, ExamResponse, ExamSubjectResponse
from app.schemas.scoring import ExamScoringConfigRequest, ScoreLineCreate

router = APIRouter(prefix="/exams", tags=["考试管理"])


@router.get("/", response_model=List[ExamResponse])
def list_exams(grade_id: str = None, exam_type: str = None, db: Session = Depends(get_db)):
    q = db.query(Exam)
    if grade_id:
        q = q.filter(Exam.grade_id == parse_uuid(grade_id))
    if exam_type:
        q = q.filter(Exam.exam_type == exam_type)
    exams = q.order_by(Exam.exam_date.desc()).all()
    result = []
    for e in exams:
        grade_name = e.grade.name if e.grade else None
        subjects = []
        for es in e.exam_subjects:
            subj = db.query(Subject).filter(Subject.id == es.subject_id).first()
            subjects.append(ExamSubjectResponse(
                id=str(es.id), subject_id=str(es.subject_id),
                subject_name=subj.name if subj else None,
                full_score=es.full_score, weight=es.weight,
                difficulty=es.difficulty, discrimination=es.discrimination,
                reliability=es.reliability,
            ))
        result.append(ExamResponse(
            id=str(e.id), name=e.name, exam_date=e.exam_date,
            exam_type=e.exam_type, grade_id=str(e.grade_id),
            grade_name=grade_name, exam_subjects=subjects,
            created_at=e.created_at,
        ))
    return result


@router.get("/{exam_id}", response_model=ExamResponse)
def get_exam(exam_id: str, db: Session = Depends(get_db)):
    e = db.query(Exam).filter(Exam.id == parse_uuid(exam_id)).first()
    if not e:
        raise HTTPException(status_code=404, detail="考试不存在")
    grade_name = e.grade.name if e.grade else None
    subjects = []
    for es in e.exam_subjects:
        subj = db.query(Subject).filter(Subject.id == es.subject_id).first()
        subjects.append(ExamSubjectResponse(
            id=str(es.id), subject_id=str(es.subject_id),
            subject_name=subj.name if subj else None,
            full_score=es.full_score, weight=es.weight,
            difficulty=es.difficulty, discrimination=es.discrimination,
            reliability=es.reliability,
        ))
    return ExamResponse(
        id=str(e.id), name=e.name, exam_date=e.exam_date,
        exam_type=e.exam_type, grade_id=str(e.grade_id),
        grade_name=grade_name, exam_subjects=subjects,
        created_at=e.created_at,
    )


@router.post("/", response_model=ExamResponse)
def create_exam(data: ExamCreate, db: Session = Depends(get_db)):
    exam = Exam(name=data.name, exam_date=data.exam_date, exam_type=data.exam_type, grade_id=parse_uuid(data.grade_id))
    db.add(exam)
    db.flush()
    for s in data.subjects:
        es = ExamSubject(
            exam_id=exam.id, subject_id=parse_uuid(s.subject_id),
            full_score=s.full_score, weight=s.weight,
            difficulty=s.difficulty, discrimination=s.discrimination,
            reliability=s.reliability,
        )
        db.add(es)
    db.commit()
    db.refresh(exam)
    grade_name = exam.grade.name if exam.grade else None
    subjects = []
    for es in exam.exam_subjects:
        subj = db.query(Subject).filter(Subject.id == es.subject_id).first()
        subjects.append(ExamSubjectResponse(
            id=str(es.id), subject_id=str(es.subject_id),
            subject_name=subj.name if subj else None,
            full_score=es.full_score, weight=es.weight,
            difficulty=es.difficulty, discrimination=es.discrimination,
            reliability=es.reliability,
        ))
    return ExamResponse(
        id=str(exam.id), name=exam.name, exam_date=exam.exam_date,
        exam_type=exam.exam_type, grade_id=str(exam.grade_id),
        grade_name=grade_name, exam_subjects=subjects,
        created_at=exam.created_at,
    )


@router.put("/{exam_id}", response_model=ExamResponse)
def update_exam(exam_id: str, data: ExamCreate, db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.id == parse_uuid(exam_id)).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    exam.name = data.name
    exam.exam_date = data.exam_date
    exam.exam_type = data.exam_type
    exam.grade_id = parse_uuid(data.grade_id)
    db.query(ExamSubject).filter(ExamSubject.exam_id == exam.id).delete()
    for s in data.subjects:
        es = ExamSubject(
            exam_id=exam.id, subject_id=parse_uuid(s.subject_id),
            full_score=s.full_score, weight=s.weight,
            difficulty=s.difficulty, discrimination=s.discrimination,
            reliability=s.reliability,
        )
        db.add(es)
    db.commit()
    db.refresh(exam)
    grade_name = exam.grade.name if exam.grade else None
    subjects = []
    for es in exam.exam_subjects:
        subj = db.query(Subject).filter(Subject.id == es.subject_id).first()
        subjects.append(ExamSubjectResponse(
            id=str(es.id), subject_id=str(es.subject_id),
            subject_name=subj.name if subj else None,
            full_score=es.full_score, weight=es.weight,
            difficulty=es.difficulty, discrimination=es.discrimination,
            reliability=es.reliability,
        ))
    return ExamResponse(
        id=str(exam.id), name=exam.name, exam_date=exam.exam_date,
        exam_type=exam.exam_type, grade_id=str(exam.grade_id),
        grade_name=grade_name, exam_subjects=subjects,
        created_at=exam.created_at,
    )


@router.delete("/{exam_id}")
def delete_exam(exam_id: str, db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.id == parse_uuid(exam_id)).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    db.delete(exam)
    db.commit()
    return {"message": "已删除"}

# ============ 赋分制：考试配置 + 分数线 ============


@router.get("/{exam_id}/scoring-config")
def get_exam_scoring_config(exam_id: str, db: Session = Depends(get_db)):
    """获取考试各科目赋分配置。"""
    from app.models.scoring import ScoringScheme
    exam = db.query(Exam).filter(Exam.id == parse_uuid(exam_id)).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    result = []
    for es in exam.exam_subjects:
        scheme = None
        if es.scheme_id:
            sc = db.query(ScoringScheme).filter(ScoringScheme.id == es.scheme_id).first()
            if sc:
                scheme = {"id": str(sc.id), "name": sc.name}
        result.append({
            "exam_subject_id": str(es.id),
            "subject_id": str(es.subject_id),
            "subject_name": es.subject.name if es.subject else "",
            "scoring_type": es.scoring_type,
            "scheme_id": str(es.scheme_id) if es.scheme_id else None,
            "scheme_name": scheme["name"] if scheme else None,
            "conversion_mode": es.conversion_mode,
        })
    return result


@router.put("/{exam_id}/scoring-config")
def update_exam_scoring_config(exam_id: str, data: ExamScoringConfigRequest, db: Session = Depends(get_db)):
    """批量配置科目赋分：scoring_type / scheme_id / conversion_mode。"""
    from app.models.scoring import ScoringScheme
    exam = db.query(Exam).filter(Exam.id == parse_uuid(exam_id)).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    es_map = {str(es.id): es for es in exam.exam_subjects}
    for cfg in data.subjects:
        es = es_map.get(cfg.exam_subject_id)
        if not es:
            raise HTTPException(status_code=400, detail=f"科目配置不属于该考试: {cfg.exam_subject_id}")
        if cfg.scoring_type not in ("raw", "converted"):
            raise HTTPException(status_code=400, detail="scoring_type 仅支持 raw / converted")
        if cfg.conversion_mode not in ("auto", "manual"):
            raise HTTPException(status_code=400, detail="conversion_mode 仅支持 auto / manual")
        if cfg.scoring_type == "converted":
            if not cfg.scheme_id:
                raise HTTPException(status_code=400, detail="赋分科目必须选择赋分方案")
            sc = db.query(ScoringScheme).filter(ScoringScheme.id == parse_uuid(cfg.scheme_id)).first()
            if not sc:
                raise HTTPException(status_code=404, detail="赋分方案不存在")
            es.scheme_id = parse_uuid(cfg.scheme_id)
        else:
            es.scheme_id = None
        es.scoring_type = cfg.scoring_type
        es.conversion_mode = cfg.conversion_mode
    db.commit()
    return {"message": "赋分配置已保存"}


@router.get("/{exam_id}/score-lines")
def list_score_lines(exam_id: str, db: Session = Depends(get_db)):
    from app.models.scoring import ScoreLine
    exam_uuid = parse_uuid(exam_id)
    exam = db.query(Exam).filter(Exam.id == exam_uuid).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    lines = db.query(ScoreLine).filter(ScoreLine.exam_id == exam_uuid).all()
    result = []
    for line in lines:
        subject_name = None
        if line.subject_id:
            subj = db.query(Subject).filter(Subject.id == line.subject_id).first()
            subject_name = subj.name if subj else None
        result.append({
            "id": str(line.id),
            "exam_id": str(line.exam_id),
            "line_name": line.line_name,
            "line_type": line.line_type,
            "subject_id": str(line.subject_id) if line.subject_id else None,
            "subject_name": subject_name,
            "score_value": line.score_value,
            "source": line.source,
        })
    return result


@router.post("/{exam_id}/score-lines")
def save_score_lines(exam_id: str, data: List[ScoreLineCreate], db: Session = Depends(get_db)):
    """新增/批量保存分数线（同线名同类型覆盖）。"""
    from app.models.scoring import ScoreLine
    items = data
    exam_uuid = parse_uuid(exam_id)
    exam = db.query(Exam).filter(Exam.id == exam_uuid).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    saved = 0
    for item in items:
        subject_uuid = None
        if item.line_type == "subject":
            if not item.subject_id:
                raise HTTPException(status_code=400, detail=f"单科线 {item.line_name} 必须指定科目")
            subject_uuid = parse_uuid(item.subject_id)
        existing = (
            db.query(ScoreLine)
            .filter(
                ScoreLine.exam_id == exam_uuid,
                ScoreLine.line_name == item.line_name,
                ScoreLine.line_type == item.line_type,
                ScoreLine.subject_id == subject_uuid,
            )
            .first()
        )
        if existing:
            existing.score_value = item.score_value
            existing.source = item.source
        else:
            db.add(ScoreLine(
                exam_id=exam_uuid,
                line_name=item.line_name,
                line_type=item.line_type,
                subject_id=subject_uuid,
                score_value=item.score_value,
                source=item.source,
            ))
        saved += 1
    db.commit()
    return {"message": f"已保存 {saved} 条分数线"}


@router.post("/{exam_id}/score-lines/import")
def import_score_lines(exam_id: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Excel 批量导入分数线：考试、线名、类型(总分/单科)、科目、分数、来源。"""
    import io
    import openpyxl
    from app.models.scoring import ScoreLine
    exam_uuid = parse_uuid(exam_id)
    exam = db.query(Exam).filter(Exam.id == exam_uuid).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_index = {h: i for i, h in enumerate(headers)}

    def col(row, name):
        idx = col_index.get(name)
        return row[idx].value if idx is not None and idx < len(row) else None

    imported = 0
    errors = []
    for row in ws.iter_rows(min_row=2):
        if all(c.value is None or str(c.value).strip() == "" for c in row):
            continue
        line_name = col(row, "线名") or col(row, "分数线名称")
        line_type = str(col(row, "类型") or "总分").strip()
        subject_name = col(row, "科目")
        score_value = col(row, "分数")
        source = str(col(row, "来源") or "official").strip()
        if not line_name or score_value is None:
            errors.append(f"第{row[0].row}行: 线名或分数为空，跳过")
            continue
        line_type = "subject" if line_type in ("单科", "单科线", "subject") else "total"
        subject_uuid = None
        if line_type == "subject":
            subj = db.query(Subject).filter(Subject.name == str(subject_name).strip()).first() if subject_name else None
            if not subj:
                errors.append(f"第{row[0].row}行: 科目 [{subject_name}] 不存在，跳过")
                continue
            subject_uuid = subj.id
        existing = (
            db.query(ScoreLine)
            .filter(
                ScoreLine.exam_id == exam_uuid,
                ScoreLine.line_name == str(line_name).strip(),
                ScoreLine.line_type == line_type,
                ScoreLine.subject_id == subject_uuid,
            )
            .first()
        )
        if existing:
            existing.score_value = float(score_value)
            existing.source = source
        else:
            db.add(ScoreLine(
                exam_id=exam_uuid,
                line_name=str(line_name).strip(),
                line_type=line_type,
                subject_id=subject_uuid,
                score_value=float(score_value),
                source=source,
            ))
        imported += 1
    db.commit()
    return {"message": f"成功导入 {imported} 条分数线", "errors": errors[:20], "error_count": len(errors)}
